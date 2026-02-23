from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
TRACK_A_SNAPSHOT_GOLDEN = (
    REPO_ROOT
    / "tests"
    / "fixtures"
    / "golden"
    / "track_a_snapshot_offline_00126380_2024.json"
)
TRACK_C_FIXTURE_DIR = REPO_ROOT / "tests" / "fixtures" / "track_c" / "basic_bundle"


def _cli_env() -> dict[str, str]:
    env = os.environ.copy()
    src_path = str(REPO_ROOT / "src")
    current_pythonpath = env.get("PYTHONPATH")
    if current_pythonpath:
        env["PYTHONPATH"] = f"{src_path}{os.pathsep}{current_pythonpath}"
    else:
        env["PYTHONPATH"] = src_path
    return env


def _run_step6_e2e(cwd: Path) -> dict[str, object]:
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "dart_pipeline.cli",
            "step6-e2e",
            "--corp-name",
            "OfflineCorp",
            "--bsns-year",
            "2024",
            "--snapshot-json",
            str(TRACK_A_SNAPSHOT_GOLDEN),
            "--xbrl-dir",
            str(TRACK_C_FIXTURE_DIR),
            "--threshold",
            "1.0",
        ],
        cwd=cwd,
        env=_cli_env(),
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0, (
        f"step6-e2e failed\nstdout:\n{result.stdout}\nstderr:\n{result.stderr}"
    )
    return json.loads(result.stdout)


def test_cli_step6_e2e_offline_golden_regression(tmp_path: Path) -> None:
    summary = _run_step6_e2e(tmp_path)

    assert summary["routing"] == {
        "route": "TRACK_B_FALLBACK",
        "reason_code": "CRITICAL_ROLE_MISSING",
        "fallback_required": True,
        "coverage_score": 0.75,
    }

    artifacts = summary["artifacts"]
    expected_relative_artifacts = {
        "track_a_snapshot_json": "out/00126380_2024_track_a_snapshot.json",
        "track_a_snapshot_report_xlsx": "out/00126380_2024_track_a_snapshot_report.xlsx",
        "track_c_route_json": "out/00126380_2024_track_c_route.json",
        "track_c_route_xlsx": "out/00126380_2024_track_c_route.xlsx",
    }
    for key, expected_relative_path in expected_relative_artifacts.items():
        assert artifacts[key] == expected_relative_path
        assert expected_relative_path.startswith("out/")
        assert (tmp_path / expected_relative_path).exists()
    assert artifacts["downloaded_xbrl_dir"] is None
    assert artifacts["xbrl_dir"] == str(TRACK_C_FIXTURE_DIR)

    route_payload = json.loads(
        (tmp_path / expected_relative_artifacts["track_c_route_json"]).read_text(
            encoding="utf-8"
        )
    )
    assert route_payload["decision"] == {
        "route": "TRACK_B_FALLBACK",
        "reason_code": "CRITICAL_ROLE_MISSING",
    }
    assert route_payload["fallback_required"] is True
    assert route_payload["report"] == {
        "required_roles": ["d822105", "d831150", "d838000", "d851100"],
        "found_roles": ["d822105", "d831150", "d838000"],
        "missing_roles": ["d851100"],
        "critical_missing_roles": ["d851100"],
        "coverage_score": 0.75,
    }

    handoff_request = route_payload["track_b_handoff_request"]
    assert handoff_request == {
        "corp_code": "00126380",
        "bsns_year": "2024",
        "reprt_code": "11011",
        "rcept_no": "20240301000001",
        "rcept_dt": "20240301",
        "fs_div": "CFS",
        "idempotency_key": "6aaeb3a5c10a02a8043c7e35ede5680226591c5b520a9f6ca2acc9c184190e7b",
        "reason_code": "CRITICAL_ROLE_MISSING",
        "missing_roles": ["d851100"],
        "critical_missing_roles": ["d851100"],
        "coverage_score": 0.75,
    }
    assert isinstance(handoff_request["idempotency_key"], str)
    assert len(handoff_request["idempotency_key"]) == 64

    workbook = load_workbook(tmp_path / expected_relative_artifacts["track_a_snapshot_report_xlsx"])
    assert workbook.sheetnames == [
        "kpi_summary",
        "metadata",
        "BS",
        "IS",
        "CIS",
        "CF",
        "SCE",
    ]

    kpi_rows = {
        row[0]: row
        for row in workbook["kpi_summary"].iter_rows(min_row=2, max_col=7, values_only=True)
    }
    assert set(kpi_rows) == {"Revenue", "OperatingIncome", "NetIncome"}

    revenue = kpi_rows["Revenue"]
    assert revenue[2] == "ifrs-full_Revenue"
    assert revenue[3] == 200
    assert revenue[4] == 180
    assert revenue[5] == pytest.approx(0.1111111111111111)
    assert revenue[6] is None

    operating_income = kpi_rows["OperatingIncome"]
    assert operating_income[2] == "dart_OperatingIncomeLoss"
    assert operating_income[3] == 40
    assert operating_income[4] == 30
    assert operating_income[5] == pytest.approx(0.3333333333333333)
    assert operating_income[6] == pytest.approx(0.2)

    net_income = kpi_rows["NetIncome"]
    assert net_income[2] == "ifrs-full_ProfitLoss"
    assert net_income[3] == 25
    assert net_income[4] == 20
    assert net_income[5] == pytest.approx(0.25)
    assert net_income[6] == pytest.approx(0.125)

    is_rows = list(workbook["IS"].iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in is_rows] == [6, 18, 23]
    assert [row[1] for row in is_rows] == [
        "dart_OperatingIncomeLoss",
        "ifrs-full_ProfitLoss",
        "ifrs-full_Revenue",
    ]
