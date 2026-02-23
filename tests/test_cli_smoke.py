from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
TRACK_C_FIXTURE_DIR = REPO_ROOT / "tests" / "fixtures" / "track_c" / "basic_bundle"


def _cli_env() -> dict[str, str]:
    env = os.environ.copy()
    src_path = str(REPO_ROOT / "src")
    current_pythonpath = env.get("PYTHONPATH")
    if current_pythonpath:
        env["PYTHONPATH"] = f"{src_path}{os.pathsep}{current_pythonpath}"
    else:
        env["PYTHONPATH"] = src_path
    return env


def _run_cli(*args: str, cwd: Path) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, "-m", "dart_pipeline.cli", *args],
        cwd=cwd,
        env=_cli_env(),
        capture_output=True,
        text=True,
        check=False,
    )


def _integration_payload(*, fallback_required: bool) -> dict[str, object]:
    if fallback_required:
        route = "TRACK_B_FALLBACK"
        reason_code = "CRITICAL_ROLE_MISSING"
        required_roles = ["d831150", "d851100"]
        missing_roles = ["d851100"]
        critical_missing_roles = ["d851100"]
        coverage_score = 0.5
    else:
        route = "TRACK_C"
        reason_code = "COVERAGE_PASS"
        required_roles = ["d831150"]
        missing_roles = []
        critical_missing_roles = []
        coverage_score = 1.0

    return {
        "track_a_snapshot": {
            "corp_code": "00126380",
            "rcept_no": "20240301000001",
            "rcept_dt": "20240301",
            "bsns_year": "2024",
            "reprt_code": "11011",
            "fs_div": "CFS",
            "rows": [
                {
                    "corp_code": "00126380",
                    "rcept_no": "20240301000001",
                    "rcept_dt": "20240301",
                    "bsns_year": "2024",
                    "reprt_code": "11011",
                    "fs_div": "CFS",
                    "sj_div": "BS",
                    "account_id": "ifrs-full_Assets",
                    "account_nm": "Assets",
                    "ord": 1,
                    "source_row_idx": 0,
                    "thstrm_amount_raw": "100",
                    "thstrm_amount": 100,
                    "frmtrm_amount_raw": "90",
                    "frmtrm_amount": 90,
                    "bfefrmtrm_amount_raw": "80",
                    "bfefrmtrm_amount": 80,
                }
            ],
        },
        "track_c_notes": [
            {
                "role_code": "D831150",
                "role_name": "Balance Sheet",
                "accounts": [],
                "members": [],
            }
        ],
        "routing_decision": {
            "route": route,
            "reason_code": reason_code,
        },
        "coverage_report": {
            "required_roles": required_roles,
            "found_roles": ["d831150"],
            "missing_roles": missing_roles,
            "critical_missing_roles": critical_missing_roles,
            "coverage_score": coverage_score,
        },
        "fallback_required": fallback_required,
    }


def _track_a_snapshot_payload() -> dict[str, object]:
    base = {
        "corp_code": "00126380",
        "rcept_no": "20240301000001",
        "rcept_dt": "20240301",
        "bsns_year": "2024",
        "reprt_code": "11011",
        "fs_div": "CFS",
    }
    rows = [
        {
            **base,
            "sj_div": "IS",
            "account_id": "ifrs-full_Revenue",
            "account_nm": "매출액",
            "ord": 23,
            "source_row_idx": 2,
            "thstrm_amount_raw": "200",
            "thstrm_amount": 200,
            "frmtrm_amount_raw": "180",
            "frmtrm_amount": 180,
            "bfefrmtrm_amount_raw": "160",
            "bfefrmtrm_amount": 160,
        },
        {
            **base,
            "sj_div": "IS",
            "account_id": "dart_OperatingIncomeLoss",
            "account_nm": "영업이익",
            "ord": 6,
            "source_row_idx": 0,
            "thstrm_amount_raw": "40",
            "thstrm_amount": 40,
            "frmtrm_amount_raw": "30",
            "frmtrm_amount": 30,
            "bfefrmtrm_amount_raw": "20",
            "bfefrmtrm_amount": 20,
        },
        {
            **base,
            "sj_div": "IS",
            "account_id": "ifrs-full_ProfitLoss",
            "account_nm": "당기순이익",
            "ord": 18,
            "source_row_idx": 1,
            "thstrm_amount_raw": "25",
            "thstrm_amount": 25,
            "frmtrm_amount_raw": "20",
            "frmtrm_amount": 20,
            "bfefrmtrm_amount_raw": "15",
            "bfefrmtrm_amount": 15,
        },
        {
            **base,
            "sj_div": "BS",
            "account_id": "ifrs-full_Assets",
            "account_nm": "자산총계",
            "ord": 7,
            "source_row_idx": 0,
            "thstrm_amount_raw": "500",
            "thstrm_amount": 500,
            "frmtrm_amount_raw": "450",
            "frmtrm_amount": 450,
            "bfefrmtrm_amount_raw": "400",
            "bfefrmtrm_amount": 400,
        },
    ]
    return {
        "corp_code": base["corp_code"],
        "rcept_no": base["rcept_no"],
        "rcept_dt": base["rcept_dt"],
        "bsns_year": base["bsns_year"],
        "reprt_code": base["reprt_code"],
        "fs_div": base["fs_div"],
        "rows": rows,
    }


def test_cli_help_returns_success() -> None:
    result = _run_cli("--help", cwd=REPO_ROOT)

    assert result.returncode == 0
    assert "tieout" in result.stdout
    assert "restatement" in result.stdout
    assert "coverage" in result.stdout
    assert "handoff-request" in result.stdout
    assert "track-c-helpers" in result.stdout
    assert "track-c-route" in result.stdout
    assert "track-a-excel" in result.stdout


@pytest.mark.parametrize("command", ["tieout", "restatement", "coverage"])
def test_cli_demo_commands_smoke(command: str, tmp_path: Path) -> None:
    result = _run_cli(command, cwd=tmp_path)

    assert result.returncode == 0
    parsed = json.loads(result.stdout)
    assert isinstance(parsed, dict)


def test_cli_default_no_write_behavior(tmp_path: Path) -> None:
    result = _run_cli("tieout", cwd=tmp_path)

    assert result.returncode == 0
    assert not (tmp_path / "out").exists()
    assert list(tmp_path.iterdir()) == []


def test_cli_output_path_under_out_writes_successfully(tmp_path: Path) -> None:
    result = _run_cli("coverage", "--output", "out/demo_coverage.json", cwd=tmp_path)

    assert result.returncode == 0
    output_file = tmp_path / "out" / "demo_coverage.json"
    assert output_file.exists()

    payload = json.loads(output_file.read_text(encoding="utf-8"))
    assert "decision" in payload
    assert "report" in payload


@pytest.mark.parametrize(
    "unsafe_output",
    [
        "../escape.json",
        "out/../escape.json",
    ],
)
def test_cli_rejects_traversal_paths(tmp_path: Path, unsafe_output: str) -> None:
    result = _run_cli("tieout", "--output", unsafe_output, cwd=tmp_path)

    assert result.returncode != 0
    assert "error:" in result.stderr.lower()
    assert "output" in result.stderr.lower()


def test_cli_rejects_absolute_path_outside_repo(tmp_path: Path) -> None:
    outside = tmp_path.parent / "outside.json"
    result = _run_cli("restatement", "--output", str(outside), cwd=tmp_path)

    assert result.returncode != 0
    assert "under ./out" in result.stderr


def test_cli_rejects_symlink_escape(tmp_path: Path) -> None:
    out_dir = tmp_path / "out"
    out_dir.mkdir()

    outside_dir = tmp_path / "outside"
    outside_dir.mkdir()

    link_path = out_dir / "link"
    try:
        link_path.symlink_to(outside_dir, target_is_directory=True)
    except OSError:
        pytest.skip("symlink creation unsupported in this test environment")

    result = _run_cli("tieout", "--output", "out/link/escaped.json", cwd=tmp_path)

    assert result.returncode != 0
    assert "under ./out" in result.stderr
    assert not (outside_dir / "escaped.json").exists()


def test_cli_handoff_request_success(tmp_path: Path) -> None:
    integration_path = tmp_path / "integration_fallback.json"
    integration_path.write_text(
        json.dumps(_integration_payload(fallback_required=True), indent=2),
        encoding="utf-8",
    )

    result = _run_cli(
        "handoff-request",
        "--integration-json",
        str(integration_path),
        cwd=tmp_path,
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["corp_code"] == "00126380"
    assert payload["bsns_year"] == "2024"
    assert payload["reason_code"] == "CRITICAL_ROLE_MISSING"
    assert payload["missing_roles"] == ["d851100"]
    assert payload["critical_missing_roles"] == ["d851100"]
    assert payload["coverage_score"] == 0.5
    assert isinstance(payload["idempotency_key"], str)
    assert len(payload["idempotency_key"]) == 64


def test_cli_handoff_request_rejects_track_c_integration(tmp_path: Path) -> None:
    integration_path = tmp_path / "integration_track_c.json"
    integration_path.write_text(
        json.dumps(_integration_payload(fallback_required=False), indent=2),
        encoding="utf-8",
    )

    result = _run_cli(
        "handoff-request",
        "--integration-json",
        str(integration_path),
        cwd=tmp_path,
    )

    assert result.returncode == 2
    assert "Track B handoff request requires TRACK_B_FALLBACK route" in result.stderr


def test_cli_track_c_helpers_success() -> None:
    result = _run_cli(
        "track-c-helpers",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        cwd=REPO_ROOT,
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload == {
        "sga_accounts": {
            "dart_DepreciationExpenseSellingGeneralAdministrativeExpenses": "감가상각비",
            "dart_SalariesWagesSellingGeneralAdministrativeExpenses": "급여",
        },
        "segment_members": [
            {
                "account_id": "entity00134477_HeadquartersMember",
                "label_ko": "본사",
                "source": "company",
            },
            {
                "account_id": "entity00134477_SalesDomesticMember",
                "label_ko": "국내",
                "source": "company",
            },
        ],
    }


def test_cli_track_c_helpers_rejects_invalid_xbrl_dir(tmp_path: Path) -> None:
    result = _run_cli(
        "track-c-helpers",
        "--xbrl-dir",
        str(tmp_path / "missing_xbrl_dir"),
        cwd=tmp_path,
    )

    assert result.returncode == 2
    assert "error:" in result.stderr
    assert "xbrl_dir" in result.stderr


def test_cli_track_c_route_success_track_c() -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--required-role",
        "D831150",
        "--required-role",
        "D838000",
        "--threshold",
        "1.0",
        cwd=REPO_ROOT,
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["decision"]["route"] == "TRACK_C"
    assert payload["decision"]["reason_code"] == "COVERAGE_PASS"
    assert payload["report"]["required_roles"] == ["d822105", "d831150", "d838000"]
    assert payload["report"]["coverage_score"] == 1.0
    assert payload["fallback_required"] is False


def test_cli_track_c_route_critical_missing_fallback() -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--required-role",
        "D831150",
        "--required-role",
        "D838000",
        "--required-role",
        "D851100",
        "--critical-role",
        "D851100",
        "--threshold",
        "1.0",
        cwd=REPO_ROOT,
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["decision"]["route"] == "TRACK_B_FALLBACK"
    assert payload["decision"]["reason_code"] == "CRITICAL_ROLE_MISSING"
    assert payload["report"]["critical_missing_roles"] == ["d851100"]
    assert payload["fallback_required"] is True


def test_cli_track_c_route_emit_handoff_request_fallback() -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--required-role",
        "D831150",
        "--required-role",
        "D838000",
        "--required-role",
        "D851100",
        "--critical-role",
        "D851100",
        "--threshold",
        "1.0",
        "--emit-handoff-request",
        "--corp-code",
        "00126380",
        "--bsns-year",
        "2024",
        "--rcept-no",
        "20240301000001",
        "--rcept-dt",
        "20240301",
        cwd=REPO_ROOT,
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    handoff_request = payload["track_b_handoff_request"]
    assert payload["decision"]["route"] == "TRACK_B_FALLBACK"
    assert handoff_request is not None
    assert isinstance(handoff_request["idempotency_key"], str)
    assert len(handoff_request["idempotency_key"]) == 64
    assert handoff_request["reason_code"] == payload["decision"]["reason_code"]
    assert handoff_request["missing_roles"] == payload["report"]["missing_roles"]
    assert (
        handoff_request["critical_missing_roles"]
        == payload["report"]["critical_missing_roles"]
    )


def test_cli_track_c_route_emit_handoff_request_track_c_is_null() -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--required-role",
        "D831150",
        "--required-role",
        "D838000",
        "--threshold",
        "1.0",
        "--emit-handoff-request",
        "--corp-code",
        "00126380",
        "--bsns-year",
        "2024",
        "--rcept-no",
        "20240301000001",
        "--rcept-dt",
        "20240301",
        cwd=REPO_ROOT,
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["decision"]["route"] == "TRACK_C"
    assert payload["track_b_handoff_request"] is None


def test_cli_track_c_route_emit_handoff_request_requires_metadata() -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--threshold",
        "1.0",
        "--emit-handoff-request",
        cwd=REPO_ROOT,
    )

    assert result.returncode == 2
    assert "--emit-handoff-request requires" in result.stderr
    assert "--corp-code" in result.stderr
    assert "--bsns-year" in result.stderr
    assert "--rcept-no" in result.stderr
    assert "--rcept-dt" in result.stderr


def test_cli_track_c_route_rejects_invalid_threshold() -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--threshold",
        "1.1",
        cwd=REPO_ROOT,
    )

    assert result.returncode == 2
    assert "threshold" in result.stderr.lower()


def test_cli_track_c_route_supports_role_alias_json(tmp_path: Path) -> None:
    alias_path = tmp_path / "role_aliases.json"
    alias_path.write_text(
        json.dumps(
            {
                "role_sga": "D831150",
                "role_ppe": "D822105",
                "role_eps": "D838000",
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "ROLE_SGA",
        "--required-role",
        "ROLE_PPE",
        "--required-role",
        "ROLE_EPS",
        "--threshold",
        "1.0",
        "--role-alias-json",
        str(alias_path),
        cwd=REPO_ROOT,
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["decision"]["route"] == "TRACK_C"
    assert payload["report"]["required_roles"] == ["d822105", "d831150", "d838000"]
    assert payload["report"]["missing_roles"] == []


def test_cli_track_c_route_rejects_invalid_role_alias_json(tmp_path: Path) -> None:
    alias_path = tmp_path / "invalid_role_aliases.json"
    alias_path.write_text(json.dumps(["not-an-object"]), encoding="utf-8")

    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--role-alias-json",
        str(alias_path),
        cwd=REPO_ROOT,
    )

    assert result.returncode == 2
    assert "role-alias-json" in result.stderr


def test_cli_track_a_excel_success_with_kpi_and_sorted_sheets(tmp_path: Path) -> None:
    snapshot_path = tmp_path / "track_a_snapshot.json"
    snapshot_path.write_text(
        json.dumps(_track_a_snapshot_payload(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    result = _run_cli(
        "track-a-excel",
        "--snapshot-json",
        str(snapshot_path),
        "--excel-output",
        "out/track_a_snapshot.xlsx",
        cwd=tmp_path,
    )

    assert result.returncode == 0
    payload = json.loads(result.stdout)
    assert payload["row_count"] == 4

    excel_path = tmp_path / "out" / "track_a_snapshot.xlsx"
    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == [
        "kpi_summary",
        "metadata",
        "BS",
        "IS",
        "CIS",
        "CF",
        "SCE",
    ]

    is_sheet = workbook["IS"]
    is_rows = [row for row in is_sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]
    assert [row[0] for row in is_rows] == [6, 18, 23]

    kpi_sheet = workbook["kpi_summary"]
    kpi_map = {
        row[0].value: row
        for row in kpi_sheet.iter_rows(min_row=2)
        if row[0].value is not None
    }
    assert kpi_map["Revenue"][3].value == 200
    assert kpi_map["OperatingIncome"][3].value == 40
    assert kpi_map["NetIncome"][3].value == 25


def test_cli_track_a_excel_rejects_invalid_snapshot_json(tmp_path: Path) -> None:
    snapshot_path = tmp_path / "invalid_snapshot.json"
    snapshot_path.write_text(json.dumps({"invalid": True}), encoding="utf-8")

    result = _run_cli(
        "track-a-excel",
        "--snapshot-json",
        str(snapshot_path),
        "--excel-output",
        "out/invalid.xlsx",
        cwd=tmp_path,
    )

    assert result.returncode == 2
    assert "snapshot-json" in result.stderr.lower()


def test_cli_track_a_excel_rejects_invalid_excel_output_path(tmp_path: Path) -> None:
    snapshot_path = tmp_path / "track_a_snapshot.json"
    snapshot_path.write_text(
        json.dumps(_track_a_snapshot_payload(), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    result = _run_cli(
        "track-a-excel",
        "--snapshot-json",
        str(snapshot_path),
        "--excel-output",
        "../escape.xlsx",
        cwd=tmp_path,
    )

    assert result.returncode == 2
    assert "excel-output" in result.stderr.lower()


def test_cli_track_c_route_writes_excel_output(tmp_path: Path) -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--required-role",
        "D831150",
        "--required-role",
        "D838000",
        "--threshold",
        "1.0",
        "--excel-output",
        "out/track_c_route.xlsx",
        cwd=tmp_path,
    )

    assert result.returncode == 0

    excel_path = tmp_path / "out" / "track_c_route.xlsx"
    assert excel_path.exists()

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == ["summary", "roles", "track_b_handoff_request"]

    summary_sheet = workbook["summary"]
    summary_values = {
        row[0].value: row[1].value
        for row in summary_sheet.iter_rows(min_row=2)
    }
    assert summary_values["route"] == "TRACK_C"
    assert summary_values["reason_code"] == "COVERAGE_PASS"
    assert summary_values["fallback_required"] is False


def test_cli_track_c_route_excel_includes_handoff_request(tmp_path: Path) -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--required-role",
        "D831150",
        "--required-role",
        "D838000",
        "--required-role",
        "D851100",
        "--critical-role",
        "D851100",
        "--threshold",
        "1.0",
        "--emit-handoff-request",
        "--corp-code",
        "00126380",
        "--bsns-year",
        "2024",
        "--rcept-no",
        "20240301000001",
        "--rcept-dt",
        "20240301",
        "--excel-output",
        "out/track_c_route_handoff.xlsx",
        cwd=tmp_path,
    )

    assert result.returncode == 0

    workbook = load_workbook(tmp_path / "out" / "track_c_route_handoff.xlsx")
    handoff_sheet = workbook["track_b_handoff_request"]
    handoff_values = {
        row[0].value: row[1].value
        for row in handoff_sheet.iter_rows(min_row=2)
    }
    assert handoff_values["reason_code"] == "CRITICAL_ROLE_MISSING"
    assert isinstance(handoff_values["idempotency_key"], str)
    assert len(handoff_values["idempotency_key"]) == 64


def test_cli_track_c_route_rejects_invalid_excel_output_path(tmp_path: Path) -> None:
    result = _run_cli(
        "track-c-route",
        "--xbrl-dir",
        str(TRACK_C_FIXTURE_DIR),
        "--required-role",
        "D822105",
        "--threshold",
        "1.0",
        "--excel-output",
        "../escape.xlsx",
        cwd=tmp_path,
    )

    assert result.returncode == 2
    assert "excel-output" in result.stderr.lower()
