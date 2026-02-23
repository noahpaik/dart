from __future__ import annotations

import argparse
import json
import math
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

from pydantic import ValidationError

from dart_pipeline.contracts import (
    Route,
    RoutingReasonCode,
    Step6TrackCIntegrationResult,
    TrackASnapshot,
)
from dart_pipeline.pipeline_step6 import build_track_b_handoff_request
from dart_pipeline.routing import route_by_coverage, route_from_track_c_roles
from dart_pipeline.track_c import (
    extract_segment_members,
    extract_sga_accounts,
    parse_xbrl_notes,
)
from dart_pipeline.timeseries import build_dual_views
from dart_pipeline.validation import run_tieout


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="dart-pipeline-cli",
        description="Run deterministic demo commands for DART pipeline components.",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    tieout = subparsers.add_parser("tieout", help="Run tie-out demo")
    tieout.add_argument(
        "--output",
        type=str,
        help="Optional JSON output path under ./out",
    )

    restatement = subparsers.add_parser("restatement", help="Run restatement demo")
    restatement.add_argument(
        "--output",
        type=str,
        help="Optional JSON output path under ./out",
    )

    coverage = subparsers.add_parser("coverage", help="Run coverage-routing demo")
    coverage.add_argument(
        "--output",
        type=str,
        help="Optional JSON output path under ./out",
    )

    handoff_request = subparsers.add_parser(
        "handoff-request",
        help="Build Track-B handoff request from Step6 integration JSON",
    )
    handoff_request.add_argument(
        "--integration-json",
        type=str,
        required=True,
        help="Path to a Step6TrackCIntegrationResult JSON file",
    )

    track_c_helpers = subparsers.add_parser(
        "track-c-helpers",
        help="Extract deterministic Track C helper outputs from an XBRL directory",
    )
    track_c_helpers.add_argument(
        "--xbrl-dir",
        type=str,
        required=True,
        help="Path to XBRL directory",
    )

    track_c_route = subparsers.add_parser(
        "track-c-route",
        help="Route deterministically from real XBRL roles",
    )
    track_c_route.add_argument(
        "--xbrl-dir",
        type=str,
        required=True,
        help="Path to XBRL directory",
    )
    track_c_route.add_argument(
        "--required-role",
        action="append",
        required=True,
        help="Required role code (repeatable, at least one)",
    )
    track_c_route.add_argument(
        "--critical-role",
        action="append",
        default=[],
        help="Critical role code (repeatable)",
    )
    track_c_route.add_argument(
        "--threshold",
        type=float,
        default=0.67,
        help="Coverage threshold in [0, 1] (default: 0.67)",
    )
    track_c_route.add_argument(
        "--role-alias-json",
        type=str,
        help="Optional JSON path for role alias map {alias: canonical_role}",
    )
    track_c_route.add_argument(
        "--emit-handoff-request",
        action="store_true",
        help="Emit one-shot Track-B handoff request payload for fallback routes",
    )
    track_c_route.add_argument(
        "--corp-code",
        type=str,
        help="Corp code metadata for --emit-handoff-request",
    )
    track_c_route.add_argument(
        "--bsns-year",
        type=str,
        help="Business year metadata for --emit-handoff-request (YYYY)",
    )
    track_c_route.add_argument(
        "--reprt-code",
        type=str,
        default="11011",
        help="Report code metadata for --emit-handoff-request (default: 11011)",
    )
    track_c_route.add_argument(
        "--rcept-no",
        type=str,
        help="Receipt number metadata for --emit-handoff-request",
    )
    track_c_route.add_argument(
        "--rcept-dt",
        type=str,
        help="Receipt date metadata for --emit-handoff-request (YYYYMMDD)",
    )
    track_c_route.add_argument(
        "--fs-div",
        type=str,
        default="CFS",
        help="FS division metadata for --emit-handoff-request (default: CFS)",
    )
    track_c_route.add_argument(
        "--excel-output",
        type=str,
        help="Optional Excel output path under ./out (.xlsx)",
    )

    track_a_excel = subparsers.add_parser(
        "track-a-excel",
        help="Build a report-friendly Track-A Excel workbook from snapshot JSON",
    )
    track_a_excel.add_argument(
        "--snapshot-json",
        type=str,
        required=True,
        help="Path to a TrackASnapshot JSON file",
    )
    track_a_excel.add_argument(
        "--excel-output",
        type=str,
        required=True,
        help="Excel output path under ./out (.xlsx)",
    )

    return parser


def _validate_threshold(threshold: float) -> float:
    if not isinstance(threshold, float) or not math.isfinite(threshold):
        raise ValueError("--threshold must be a finite float in [0, 1]")
    if threshold < 0.0 or threshold > 1.0:
        raise ValueError("--threshold must be within [0, 1]")
    return threshold


def _load_role_aliases(role_alias_json_path: str | None) -> dict[str, str] | None:
    if role_alias_json_path is None:
        return None

    alias_path = Path(role_alias_json_path)
    try:
        raw_payload = alias_path.read_text(encoding="utf-8")
    except OSError as exc:
        raise ValueError(
            f"unable to read --role-alias-json at {alias_path}: {exc}"
        ) from exc

    try:
        payload = json.loads(raw_payload)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"--role-alias-json is not valid JSON ({exc.msg})"
        ) from exc

    if not isinstance(payload, dict):
        raise ValueError("--role-alias-json must be a JSON object mapping str to str")

    aliases: dict[str, str] = {}
    for raw_alias, raw_canonical in payload.items():
        if not isinstance(raw_alias, str) or not isinstance(raw_canonical, str):
            raise ValueError("--role-alias-json must map str keys to str values")

        alias = raw_alias.strip()
        canonical = raw_canonical.strip()
        if not alias or not canonical:
            raise ValueError("--role-alias-json keys and values must be non-empty strings")

        aliases[alias] = canonical

    return aliases


def _require_non_empty_metadata_arg(value: str | None, argument_name: str) -> str:
    if value is None:
        raise ValueError(f"{argument_name} is required")
    stripped = value.strip()
    if not stripped:
        raise ValueError(f"{argument_name} must be a non-empty string")
    return stripped


def _validate_emit_handoff_metadata(
    *,
    corp_code: str | None,
    bsns_year: str | None,
    reprt_code: str,
    rcept_no: str | None,
    rcept_dt: str | None,
    fs_div: str,
) -> dict[str, str]:
    missing_args: list[str] = []
    if corp_code is None:
        missing_args.append("--corp-code")
    if bsns_year is None:
        missing_args.append("--bsns-year")
    if rcept_no is None:
        missing_args.append("--rcept-no")
    if rcept_dt is None:
        missing_args.append("--rcept-dt")
    if missing_args:
        raise ValueError(
            "--emit-handoff-request requires "
            + ", ".join(missing_args)
        )

    validated_corp_code = _require_non_empty_metadata_arg(corp_code, "--corp-code")
    validated_bsns_year = _require_non_empty_metadata_arg(bsns_year, "--bsns-year")
    validated_reprt_code = _require_non_empty_metadata_arg(reprt_code, "--reprt-code")
    validated_rcept_no = _require_non_empty_metadata_arg(rcept_no, "--rcept-no")
    validated_rcept_dt = _require_non_empty_metadata_arg(rcept_dt, "--rcept-dt")
    validated_fs_div = _require_non_empty_metadata_arg(fs_div, "--fs-div")

    if len(validated_bsns_year) != 4 or not validated_bsns_year.isdigit():
        raise ValueError("--bsns-year must be a 4-digit string")
    if len(validated_rcept_dt) != 8 or not validated_rcept_dt.isdigit():
        raise ValueError("--rcept-dt must be an 8-digit YYYYMMDD string")

    return {
        "corp_code": validated_corp_code,
        "bsns_year": validated_bsns_year,
        "reprt_code": validated_reprt_code,
        "rcept_no": validated_rcept_no,
        "rcept_dt": validated_rcept_dt,
        "fs_div": validated_fs_div,
    }


def _demo_tieout() -> dict[str, Any]:
    expected = [
        {
            "metric": "revenue",
            "period": "2024Q4",
            "unit": "KRW",
            "dimensions": {"segment": "A"},
            "value": 100.0,
        }
    ]
    observed = [
        {
            "metric": "revenue",
            "period": "2024Q4",
            "unit": "KRW",
            "dimensions": {"segment": "A"},
            "value": 100.2,
        }
    ]

    result = run_tieout(
        expected=expected,
        observed=observed,
        abs_tol=0.1,
        rel_tol=0.0,
        warn_multiplier=3.0,
    )
    return result.model_dump(mode="json")


def _demo_restatement() -> dict[str, Any]:
    reports = [
        {
            "metric": "revenue",
            "period": "2024Q4",
            "unit": "KRW",
            "dimensions": {},
            "value": 100.0,
            "filing_datetime_utc": "2024-03-01T00:00:00Z",
            "rcept_no": "20240301000001",
            "source_row_idx": 0,
        },
        {
            "metric": "revenue",
            "period": "2024Q4",
            "unit": "KRW",
            "dimensions": {},
            "value": 110.0,
            "filing_datetime_utc": "2024-04-01T00:00:00Z",
            "rcept_no": "20240401000001",
            "source_row_idx": 0,
        },
    ]

    as_reported, as_latest = build_dual_views(reports)
    return {
        "as_reported": as_reported.model_dump(mode="json"),
        "as_latest": as_latest.model_dump(mode="json"),
    }


def _demo_coverage() -> dict[str, Any]:
    decision, report = route_by_coverage(
        required_roles=["income_statement", "balance_sheet", "cash_flow"],
        found_roles=["income_statement", "balance_sheet"],
        critical_roles=["cash_flow"],
        threshold=0.8,
    )
    return {
        "decision": decision.model_dump(mode="json"),
        "report": report.model_dump(mode="json") if report is not None else None,
    }


def _validate_output_path(
    raw_output: str,
    cwd: Path,
    argument_name: str = "--output",
) -> Path:
    output = Path(raw_output)
    if any(part == ".." for part in output.parts):
        raise ValueError(
            f"{argument_name} must not contain '..' path traversal segments"
        )

    out_root = cwd / "out"
    if out_root.is_symlink():
        raise ValueError("./out must not be a symlink")

    target = output if output.is_absolute() else cwd / output
    resolved_target = target.resolve(strict=False)
    resolved_out_root = out_root.resolve(strict=False)

    try:
        resolved_target.relative_to(resolved_out_root)
    except ValueError as exc:
        raise ValueError(f"{argument_name} must resolve to a path under ./out") from exc

    if resolved_target == resolved_out_root:
        raise ValueError(f"{argument_name} must point to a file path under ./out")

    return resolved_target


def _run_command(command: str) -> dict[str, Any]:
    if command == "tieout":
        return _demo_tieout()
    if command == "restatement":
        return _demo_restatement()
    if command == "coverage":
        return _demo_coverage()
    raise ValueError(f"unsupported command: {command}")


def _build_handoff_request_payload(integration_json_path: str) -> dict[str, Any]:
    integration_path = Path(integration_json_path)
    try:
        payload = integration_path.read_text(encoding="utf-8")
    except OSError as exc:
        raise ValueError(
            f"unable to read --integration-json at {integration_path}: {exc}"
        ) from exc

    try:
        integration_result = Step6TrackCIntegrationResult.model_validate_json(payload)
    except ValidationError as exc:
        first_error = exc.errors()[0] if exc.errors() else {"loc": (), "msg": str(exc)}
        location = ".".join(str(part) for part in first_error.get("loc", ()))
        message = first_error.get("msg", "invalid payload")
        detail = (
            f"{location}: {message}"
            if location
            else message
        )
        raise ValueError(
            "--integration-json is not a valid Step6TrackCIntegrationResult "
            f"({detail})"
        ) from exc

    return build_track_b_handoff_request(
        integration_result=integration_result
    ).model_dump(mode="json")


def _build_track_c_helpers_payload(xbrl_dir: str) -> dict[str, Any]:
    notes = parse_xbrl_notes(xbrl_dir=xbrl_dir)
    return {
        "sga_accounts": extract_sga_accounts(notes),
        "segment_members": [
            member.model_dump(mode="json") for member in extract_segment_members(notes)
        ],
    }


def _build_track_c_route_payload(
    *,
    xbrl_dir: str,
    required_roles: list[str],
    critical_roles: list[str],
    threshold: float,
    role_alias_json_path: str | None,
    emit_handoff_request: bool,
    corp_code: str | None,
    bsns_year: str | None,
    reprt_code: str,
    rcept_no: str | None,
    rcept_dt: str | None,
    fs_div: str,
) -> dict[str, Any]:
    threshold = _validate_threshold(threshold)
    role_aliases = _load_role_aliases(role_alias_json_path)
    notes = parse_xbrl_notes(xbrl_dir=xbrl_dir)
    decision, report = route_from_track_c_roles(
        parsed_notes=notes,
        required_roles=required_roles,
        critical_roles=critical_roles,
        threshold=threshold,
        role_aliases=role_aliases,
    )
    if decision.reason_code == RoutingReasonCode.INVALID_INPUT:
        raise ValueError(
            "invalid routing input for track-c-route "
            "(check --required-role, --critical-role, and --threshold)"
        )
    payload: dict[str, Any] = {
        "decision": decision.model_dump(mode="json"),
        "report": report.model_dump(mode="json") if report is not None else None,
        "fallback_required": decision.route == Route.TRACK_B_FALLBACK,
    }
    if not emit_handoff_request:
        return payload

    metadata = _validate_emit_handoff_metadata(
        corp_code=corp_code,
        bsns_year=bsns_year,
        reprt_code=reprt_code,
        rcept_no=rcept_no,
        rcept_dt=rcept_dt,
        fs_div=fs_div,
    )
    if decision.route == Route.TRACK_C:
        payload["track_b_handoff_request"] = None
        return payload

    integration_result = Step6TrackCIntegrationResult(
        track_a_snapshot=TrackASnapshot(
            corp_code=metadata["corp_code"],
            rcept_no=metadata["rcept_no"],
            rcept_dt=metadata["rcept_dt"],
            bsns_year=metadata["bsns_year"],
            reprt_code=metadata["reprt_code"],
            fs_div=metadata["fs_div"],
            rows=[],
        ),
        track_c_notes=[],
        routing_decision=decision,
        coverage_report=report,
        fallback_required=True,
    )
    payload["track_b_handoff_request"] = build_track_b_handoff_request(
        integration_result=integration_result
    ).model_dump(mode="json")
    return payload


def _validate_excel_output_path(raw_output: str, cwd: Path) -> Path:
    output_path = _validate_output_path(
        raw_output,
        cwd=cwd,
        argument_name="--excel-output",
    )
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError("--excel-output must end with .xlsx")
    return output_path


def _to_excel_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _write_track_c_route_excel(payload: dict[str, Any], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise ValueError(
            "openpyxl is required for --excel-output. Install dependencies first."
        ) from exc

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["field", "value"])

    decision = payload.get("decision")
    report = payload.get("report")
    summary_sheet.append([
        "route",
        decision.get("route") if isinstance(decision, dict) else None,
    ])
    summary_sheet.append([
        "reason_code",
        decision.get("reason_code") if isinstance(decision, dict) else None,
    ])
    summary_sheet.append(["fallback_required", payload.get("fallback_required")])
    summary_sheet.append([
        "coverage_score",
        report.get("coverage_score") if isinstance(report, dict) else None,
    ])

    roles_sheet = workbook.create_sheet("roles")
    roles_sheet.append(["group", "role"])
    if isinstance(report, dict):
        for group_name in (
            "required_roles",
            "found_roles",
            "missing_roles",
            "critical_missing_roles",
        ):
            group_values = report.get(group_name)
            if not isinstance(group_values, list):
                continue
            for role in group_values:
                roles_sheet.append([group_name, _to_excel_cell_value(role)])

    handoff_sheet = workbook.create_sheet("track_b_handoff_request")
    handoff_sheet.append(["field", "value"])
    handoff_payload = payload.get("track_b_handoff_request")
    if isinstance(handoff_payload, dict):
        for key in sorted(handoff_payload):
            handoff_sheet.append([key, _to_excel_cell_value(handoff_payload[key])])
    else:
        handoff_sheet.append(["value", None])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


_TRACK_A_SHEET_ORDER = ("BS", "IS", "CIS", "CF", "SCE")
_TRACK_A_EXPORT_COLUMNS = (
    "ord",
    "account_id",
    "account_nm",
    "thstrm_amount",
    "frmtrm_amount",
    "bfefrmtrm_amount",
    "thstrm_amount_raw",
    "frmtrm_amount_raw",
    "bfefrmtrm_amount_raw",
    "source_row_idx",
)


def _to_excel_numeric(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    if isinstance(value, Decimal):
        return float(value)
    return None


def _safe_ratio(numerator: int | float | None, denominator: int | float | None) -> float | None:
    if numerator is None or denominator is None:
        return None
    if denominator == 0:
        return None
    return float(numerator) / float(denominator)


def _safe_growth(current: int | float | None, prior: int | float | None) -> float | None:
    if current is None or prior is None:
        return None
    if prior == 0:
        return None
    return (float(current) - float(prior)) / abs(float(prior))


def _find_is_row(snapshot: TrackASnapshot, account_ids: tuple[str, ...]) -> Any:
    is_rows = [row for row in snapshot.rows if row.sj_div == "IS"]
    for account_id in account_ids:
        for row in is_rows:
            if row.account_id == account_id:
                return row
    return None


def _build_kpi_rows(snapshot: TrackASnapshot) -> list[dict[str, Any]]:
    revenue_row = _find_is_row(snapshot, ("ifrs-full_Revenue", "dart_Revenue"))
    operating_income_row = _find_is_row(
        snapshot,
        (
            "dart_OperatingIncomeLoss",
            "ifrs-full_ProfitLossFromOperatingActivities",
        ),
    )
    net_income_row = _find_is_row(
        snapshot,
        (
            "ifrs-full_ProfitLoss",
            "ifrs-full_ProfitLossAttributableToOwnersOfParent",
        ),
    )

    revenue_current = _to_excel_numeric(
        None if revenue_row is None else revenue_row.thstrm_amount
    )
    revenue_prior = _to_excel_numeric(
        None if revenue_row is None else revenue_row.frmtrm_amount
    )
    operating_current = _to_excel_numeric(
        None if operating_income_row is None else operating_income_row.thstrm_amount
    )
    operating_prior = _to_excel_numeric(
        None if operating_income_row is None else operating_income_row.frmtrm_amount
    )
    net_current = _to_excel_numeric(
        None if net_income_row is None else net_income_row.thstrm_amount
    )
    net_prior = _to_excel_numeric(
        None if net_income_row is None else net_income_row.frmtrm_amount
    )

    return [
        {
            "metric": "Revenue",
            "label_ko": "매출액",
            "account_id": None if revenue_row is None else revenue_row.account_id,
            "current": revenue_current,
            "prior": revenue_prior,
            "yoy": _safe_growth(revenue_current, revenue_prior),
            "margin": None,
        },
        {
            "metric": "OperatingIncome",
            "label_ko": "영업이익",
            "account_id": (
                None if operating_income_row is None else operating_income_row.account_id
            ),
            "current": operating_current,
            "prior": operating_prior,
            "yoy": _safe_growth(operating_current, operating_prior),
            "margin": _safe_ratio(operating_current, revenue_current),
        },
        {
            "metric": "NetIncome",
            "label_ko": "당기순이익",
            "account_id": None if net_income_row is None else net_income_row.account_id,
            "current": net_current,
            "prior": net_prior,
            "yoy": _safe_growth(net_current, net_prior),
            "margin": _safe_ratio(net_current, revenue_current),
        },
    ]


def _build_track_a_excel_payload(snapshot_json_path: str, excel_output_path: str, cwd: Path) -> dict[str, Any]:
    snapshot_path = Path(snapshot_json_path)
    try:
        raw_json = snapshot_path.read_text(encoding="utf-8")
    except OSError as exc:
        raise ValueError(f"unable to read --snapshot-json at {snapshot_path}: {exc}") from exc

    try:
        snapshot = TrackASnapshot.model_validate_json(raw_json)
    except ValidationError as exc:
        first_error = exc.errors()[0] if exc.errors() else {"loc": (), "msg": str(exc)}
        location = ".".join(str(part) for part in first_error.get("loc", ()))
        message = first_error.get("msg", "invalid payload")
        detail = f"{location}: {message}" if location else message
        raise ValueError(
            "--snapshot-json is not a valid TrackASnapshot "
            f"({detail})"
        ) from exc

    output_path = _validate_excel_output_path(excel_output_path, cwd=cwd)

    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise ValueError(
            "openpyxl is required for --excel-output. Install dependencies first."
        ) from exc

    workbook = Workbook()

    kpi_sheet = workbook.active
    kpi_sheet.title = "kpi_summary"
    kpi_sheet.append(["metric", "label_ko", "account_id", "current", "prior", "yoy", "margin"])
    for row in _build_kpi_rows(snapshot):
        kpi_sheet.append([
            row["metric"],
            row["label_ko"],
            row["account_id"],
            row["current"],
            row["prior"],
            row["yoy"],
            row["margin"],
        ])

    metadata_sheet = workbook.create_sheet("metadata")
    metadata_sheet.append(["field", "value"])
    for field_name in (
        "corp_code",
        "rcept_no",
        "rcept_dt",
        "bsns_year",
        "reprt_code",
        "fs_div",
    ):
        metadata_sheet.append([field_name, getattr(snapshot, field_name)])
    metadata_sheet.append(["row_count", len(snapshot.rows)])

    rows_by_sj_div: dict[str, list[Any]] = {sj_div: [] for sj_div in _TRACK_A_SHEET_ORDER}
    for row in snapshot.rows:
        rows_by_sj_div.setdefault(row.sj_div, []).append(row)

    for sj_div in _TRACK_A_SHEET_ORDER:
        sheet = workbook.create_sheet(sj_div)
        sheet.append(list(_TRACK_A_EXPORT_COLUMNS))
        sorted_rows = sorted(
            rows_by_sj_div.get(sj_div, []),
            key=lambda row: (
                row.ord,
                row.account_id,
                row.account_nm,
                row.source_row_idx,
            ),
        )
        for row in sorted_rows:
            sheet.append([
                row.ord,
                row.account_id,
                row.account_nm,
                _to_excel_numeric(row.thstrm_amount),
                _to_excel_numeric(row.frmtrm_amount),
                _to_excel_numeric(row.bfefrmtrm_amount),
                row.thstrm_amount_raw,
                row.frmtrm_amount_raw,
                row.bfefrmtrm_amount_raw,
                row.source_row_idx,
            ])

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:J{max(sheet.max_row, 1)}"
        for amount_col in ("D", "E", "F"):
            for cell in sheet[f"{amount_col}2":f"{amount_col}{sheet.max_row}"]:
                for item in cell:
                    item.number_format = "#,##0"

    kpi_sheet.freeze_panes = "A2"
    kpi_sheet.auto_filter.ref = f"A1:G{max(kpi_sheet.max_row, 1)}"
    for ratio_col in ("F", "G"):
        for cell in kpi_sheet[f"{ratio_col}2":f"{ratio_col}{kpi_sheet.max_row}"]:
            for item in cell:
                item.number_format = "0.00%"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "snapshot_json": str(snapshot_path),
        "excel_output": str(output_path),
        "sheet_names": workbook.sheetnames,
        "row_count": len(snapshot.rows),
    }


def main(argv: Sequence[str] | None = None) -> int:
    parser = _build_parser()
    args = parser.parse_args(argv)

    if args.command == "handoff-request":
        try:
            payload = _build_handoff_request_payload(args.integration_json)
        except ValueError as exc:
            parser.exit(status=2, message=f"error: {exc}\n")
    elif args.command == "track-c-helpers":
        try:
            payload = _build_track_c_helpers_payload(args.xbrl_dir)
        except ValueError as exc:
            parser.exit(status=2, message=f"error: {exc}\n")
    elif args.command == "track-c-route":
        try:
            payload = _build_track_c_route_payload(
                xbrl_dir=args.xbrl_dir,
                required_roles=args.required_role,
                critical_roles=args.critical_role,
                threshold=args.threshold,
                role_alias_json_path=args.role_alias_json,
                emit_handoff_request=args.emit_handoff_request,
                corp_code=args.corp_code,
                bsns_year=args.bsns_year,
                reprt_code=args.reprt_code,
                rcept_no=args.rcept_no,
                rcept_dt=args.rcept_dt,
                fs_div=args.fs_div,
            )
        except ValueError as exc:
            parser.exit(status=2, message=f"error: {exc}\n")
    elif args.command == "track-a-excel":
        try:
            payload = _build_track_a_excel_payload(
                args.snapshot_json,
                args.excel_output,
                cwd=Path.cwd().resolve(),
            )
        except ValueError as exc:
            parser.exit(status=2, message=f"error: {exc}\n")
    else:
        payload = _run_command(args.command)

    rendered = json.dumps(payload, sort_keys=True, indent=2) + "\n"

    cwd = Path.cwd().resolve()

    output = getattr(args, "output", None)
    if output:
        try:
            output_path = _validate_output_path(
                output,
                cwd=cwd,
                argument_name="--output",
            )
        except ValueError as exc:
            parser.exit(status=2, message=f"error: {exc}\n")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(rendered, encoding="utf-8")

    excel_output = getattr(args, "excel_output", None)
    if excel_output and args.command == "track-c-route":
        try:
            excel_output_path = _validate_excel_output_path(excel_output, cwd=cwd)
            _write_track_c_route_excel(payload, excel_output_path)
        except ValueError as exc:
            parser.exit(status=2, message=f"error: {exc}\n")

    print(rendered, end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
